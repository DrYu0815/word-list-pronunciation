import openpyxl
import requests
from tqdm import tqdm
import time
from openpyxl.styles import Font

# 设置Excel文件路径和API的URL
excel_path = 'F:\\pycharm\\pycharm\\PyCharm 2023.3.3\\project\\ieltwordlist\\雅思真题中的必背词总结.xlsx'
api_url = 'https://api.dictionaryapi.dev/api/v2/entries/en/'

# 加载Excel工作簿
wb = openpyxl.load_workbook(excel_path)
sheet = wb.active

# 检测D列注释到的最后一行
last_annotated_row = 1
for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=4).value is not None:
        last_annotated_row = i
start_row = last_annotated_row + 1  # 从下一行开始注释

print(f"上次注释的最后一行是第{last_annotated_row}行。")

# 准备进度条
pbar = tqdm(total=sheet.max_row - start_row + 1)

# 批量获取发音并写入Excel
try:
    batch_size = 10  # 每批处理的单词数量
    for row in range(start_row, sheet.max_row + 1, batch_size):
        words = [sheet.cell(row=r, column=2).value for r in range(row, min(row + batch_size, sheet.max_row + 1))]
        pronunciations = {}
        for word in words:
            if word:
                response = requests.get(api_url + word)
                if response.status_code == 200:
                    data = response.json()
                    if data[0].get('phonetics') and any('text' in phonetic for phonetic in data[0]['phonetics']):
                        pronunciation = next(phonetic['text'] for phonetic in data[0]['phonetics'] if 'text' in phonetic)
                        pronunciations[word] = pronunciation
                    else:
                        pronunciations[word] = '发音不可用'
                else:
                    print(f"API请求失败，状态码：{response.status_code}")
                time.sleep(1)  # 为了防止API限制，可以适当增加延时

        # 更新Excel文件
        for r in range(row, min(row + batch_size, sheet.max_row + 1)):
            word = sheet.cell(row=r, column=2).value
            if word in pronunciations:
                pronunciation = pronunciations[word]
                if pronunciation != '发音不可用':
                    sheet.cell(row=r, column=4, value=pronunciation)
                else:
                    # 发音未找到，标记该行为红色
                    for col in range(1, 4):
                        cell = sheet.cell(row=r, column=col)
                        cell.font = Font(color="FF0000")
            pbar.update(1)

except KeyboardInterrupt:
    # 如果用户中断执行，保存当前进度
    print("\n操作被用户中断，保存当前进度...")
finally:
    # 保存Excel文件
    wb.save(excel_path)
    pbar.close()
    completed = pbar.n
    print(f"已完成{completed}个单词的发音注释，总共{sheet.max_row}个。这次注释到了第{last_annotated_row + completed}行，总共注释了{completed}个单词。预计剩余时间：{((sheet.max_row - last_annotated_row - completed) * 1) / 60}分钟。")

# 运行脚本后，如果中断再次运行会从上次中断的地方继续
